from kivy.core.text import markup
from openpyxl import load_workbook, Workbook
from kivy.app import App
from kivy.properties import ObjectProperty
from kivy.lang import Builder
from kivy.uix.screenmanager import Screen, ScreenManager, FadeTransition
from sys import exit
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView

class Window1(Screen):
    # def popme(self):
    #     self.pop = Popup(
    #         title = "Required.",
    #         size_hint = (0.4, 0.4),
    #         content =Label(
    #             text = "Enter the filename: ",
    #             font_size = 13,   
    #         )
    #     )
    def choose(self):
        pass

    
    def onClick(self):
        exit()
    pass

class Window2(Screen):
    def onpress(self):
        pfilename = ObjectProperty(None)
        name = str(self.pfilename.text)
        try:
            self.pop = Popup(
                title = "File creation successful ..",
                size_hint = (0.4, 0.4),
                content = Label(
                    font_size = 13,
                    text = (f"The file was successfully created as {name}.xlsx !!!")
                )
            )
            wb = Workbook()
            wb.create_sheet("Purchase Book")
            sheet1 = wb["Purchase Book"]
            sheet1.append(["Date", "Invoice No", "Supplier Name", "Supplier Pan No","Item Name",
                "Quantity", "Total Purchase", "Exempt Value", "Taxable Purchase",
                "VAT", "Taxable Import", "VAT", "Capital Purchase", "VAT"])
            wb.save(name+".xlsx")
            self.pop.open()
        except Exception as e:
            self.pop = Popup(
                title = "Warning !!",
                size_hint = (0.4, 0.4),
                background_color = "red",
                content = Label(
                    font_size = 13,
                    text = "Please Enter the filename .."
                )
            )
            self.pop.open()
 
    def add(self):
        try:
            pdate =  ObjectProperty(None)
            pinvoice_no =  ObjectProperty(None)
            psupplier_name =  ObjectProperty(None)
            psupplier_pan_no =  ObjectProperty(None)
            pitem_name =  ObjectProperty(None)
            pquantity =  ObjectProperty(None)
            ptotal_purchase =  ObjectProperty(None)
            pexempt_value =  ObjectProperty(None)
            ptaxable_purchase =  ObjectProperty(None)
            ptaxable_import =  ObjectProperty(None)
            pcapital_purchase =  ObjectProperty(None)
            pfilename =  ObjectProperty(None)

            date = str(self.pdate.text)
            invoice_no = str(self.pinvoice_no.text) 
            supplier_name = str(self.psupplier_name.text)
            supplier_pan_no  = str(self.psupplier_pan_no.text)
            item_name = str(self.pitem_name.text)
            quantity = str(self.pquantity.text)
            total_purchase = str(self.ptotal_purchase.text)
            exempt_value = str(self.pexempt_value.text)
            taxable_purchase = int(self.ptaxable_purchase.text)
            vat1 = float(taxable_purchase * 0.13)
            taxable_import = int(self.ptaxable_import.text)
            vat2 = float(taxable_import * 0.13)
            capital_purchase = int(self.pcapital_purchase.text)
            vat3 = float(capital_purchase * 0.13)
            filename = str(self.pfilename.text)

        

            wb = load_workbook(filename+".xlsx")
            sheet1 = wb["Purchase Book"]
            sheet1.append([date, invoice_no, supplier_name,supplier_pan_no , item_name, quantity, total_purchase, exempt_value, taxable_purchase, str(vat1),taxable_import ,str(vat2), capital_purchase, str(vat3)])
            self.pdate.text = ""
            self.pinvoice_no.text = ""
            self.psupplier_name.text = ""
            self.psupplier_pan_no.text = ""
            self.pitem_name.text = ""         
            self.pquantity.text = ""
            self.ptotal_purchase.text = ""
            self.pexempt_value.text = ""
            self.ptaxable_purchase.text = ""
            self.ptaxable_import.text = ""
            self.pcapital_purchase.text = ""

            wb.save(filename+".xlsx")

        except Exception as e:
            self.pop = Popup(
                title = "Filename Error !!", 
                size_hint = (0.4, 0.4),
                backbround_color = "red",
                content = Label(
                    text = "Please check the Filename or Sheetname"
                )
            )
        print(date)
        pass
    pass

class Window3(Screen):
    def onpress(self):
        sfilename = ObjectProperty(None)
        name = str(self.sfilename.text)
        try:
            self.pop = Popup(
                title = "File creation successful ..",
                size_hint = (0.4, 0.4),
                content = Label(
                    font_size = 13,
                    text = (f"The file was successfully created as {name}.xlsx !!!")
                )
            )
            wb = Workbook()
            wb.create_sheet("Purchase Book")
            wb.create_sheet("Sales Book")
            sheet2 = wb["Sales Book"]
            sheet2.append(["Date", "Invoice No", "Supplier's Name", "Pan No",
                "Item Name", "Quantity", "Total Sales", "Exempt Sales", "Taxable Sales", "VAT"])
            wb.save(name+".xlsx")
            self.pop.open()

        except Exception as e:
            self.pop = Popup(
                title = "Warning !!",
                size_hint = (0.4, 0.4),
                background_color = "red",
                content = Label(
                    font_size = 13,
                    text = "Please Enter the filename .."
                )
            )
            self.pop.open()

    def add(self):
        try:
            sdate = ObjectProperty(None)
            sinvoice_no = ObjectProperty(None)
            ssupplier_name = ObjectProperty(None)
            span_no = ObjectProperty(None)
            sitem_name = ObjectProperty(None)
            squantity = ObjectProperty(None)
            stotal_sales = ObjectProperty(None)
            sexempt_sales = ObjectProperty(None)
            staxable_sales = ObjectProperty(None)
            sfilename = ObjectProperty(None)

            date = str(self.sdate.text)
            invoice_no = str(self.sinvoice_no.text)
            supplier_name = str(self.ssupplier_name.text)
            pan_no = str(self.span_no.text)
            item_name = str(self.sitem_name.text)
            quantity = str(self.squantity.text)
            total_sales = str(self.stotal_sales.text)
            exempt_sales = str(self.sexempt_sales.text)
            taxable_sales = int(self.staxable_sales.text)
            vat = float(taxable_sales * 0.13)
            filename = str(self.sfilename.text)

            wb = load_workbook(filename+".xlsx")
            sheet2 = wb["Sales Book"]
            sheet2.append([date, invoice_no,supplier_name,pan_no,item_name,quantity,total_sales,exempt_sales,str(taxable_sales),str(vat)])
            self.sdate.text = ""
            self.sinvoice_no.text = ""
            self.ssupplier_name.text = ""
            self.span_no.text = ""
            self.sitem_name.text = ""
            self.squantity.text = ""
            self.stotal_sales.text = ""
            self.sexempt_sales.text = ""
            self.staxable_sales.text = ""

            wb.save(filename+".xlsx")

        except Exception as e:
            self.pop = Popup(
                title = "Filename Error !!", 
                size_hint = (0.4, 0.4),
                backbround_color = "red",
                content = Label(
                    text = "Please check the Filename or Sheetname"
                )
            )
            pass
    pass
class  Window4(Screen):
    def create(self):
        try:
            c_text = ObjectProperty(None)
            filename = str(self.c_text.text)

            wb = Workbook()
            wb.create_sheet("Purchase Book")
            wb.create_sheet("Sales Book")
            wb.create_sheet("Stock Book")
            sheet1 = wb['Purchase Book']
            sheet2 = wb['Sales Book']
            # sheet3 = wb['Stock Book']
            sheet1.append(["Date", "Invoice No", "Supplier Name", "Supplier Pan No","Item Name",
                    "Quantity", "Total Purchase", "Exempt Value", "Taxable Purchase",
                    "VAT", "Taxable Import", "VAT", "Capital Purchase", "VAT"])
            sheet2.append(["Date", "Invoice No", "Supplier's Name", "Pan No",
                    "Item Name", "Quantity", "Total Sales", "Exempt Sales", "Taxable Sales", "VAT"])
            wb.save(filename+".xlsx")
            self.pop = Popup(
                    title = "Successful",
                    size_hint = (0.4, 0.4),
                    background_color = "white",
                    content = Label(
                        markup = True,
                        font_size = 14,
                        text = (f"[b]{filename}'s[/b] [i][b]format has been created successfully[/i][/b]")
                        ,color = "white"
                    )
                )
            self.pop.open()

        except Exception as e:
            self.pop = Popup(
                title = "Filename Error !!", 
                size_hint = (0.4, 0.4),
                backbround_color = "red",
                content = Label(
                    text = "Please check the Filename or Sheetname"
                )
            )

    pass

class WindowManager(ScreenManager):
    pass

kv = Builder.load_file("layout.kv")

class Harsita_Consultancy(App):
    def build(self):
        return kv

Harsita_Consultancy().run()